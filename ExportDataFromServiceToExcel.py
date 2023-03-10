import arcpy
import openpyxl

# Get the current date and time as a string with specific formatting, and assign it to a variable so it can be appended to the output file name
# Formatting help: https://docs.python.org/3/library/datetime.html#strftime-and-strptime-behavior
from time import strftime
currentDateAndTime = strftime("%m%d%Y_%H%M%p")

# Create a variable that breaks up the output file path and concatenates the date and time between the file name and file type
outpath = r"C:\Users\JohnsonN35\project_path\export_" + currentDateAndTime + ".xlsx"

# Convert table to Excel
# "ALIAS" means the output will have alias field names, and "DESCRIPTION" means that coded values will appear as domain descriptions rather than codes
arcpy.conversion.TableToExcel("https://services3.arcgis.com/X0xdaFqVSAx896l1/ArcGIS/rest/services/ServiceLine_Public/FeatureServer/0", outpath, "ALIAS", "DESCRIPTION")

# Load the workbook you just created 
# https://openpyxl.readthedocs.io/en/stable/usage.html
from openpyxl import load_workbook
wb = load_workbook(outpath)
ws = wb.active # Refer to the first sheet in the file (what "active" defaults to)

# OPTIONAL: Delete two columns starting at index 1
# https://openpyxl.readthedocs.io/en/latest/editing_worksheets.html
ws.delete_cols(1, 2)

# OPTIONAL: Make the spreadsheet more readable by adjusting the column widths
# https://stackoverflow.com/a/35790441
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
for col, value in dims.items():
    # Sometimes field headings are longer than the cell values, so this provides some extra space
    ws.column_dimensions[col].width = value * 1.4
    
# OPTIONAL: Freeze rows above cell A2 (could freeze first column in addition by choosing 'B2' instead)
freezeRow = ws['A2']
ws.freeze_panes = freezeRow

# Re-save the file
wb.save(outpath)


